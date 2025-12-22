# interfaz_programa.py — v9.1 (Corrección detección Word y Formato Premios)
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import sqlite3, os, re, sys, traceback
from pathlib import Path
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins

# --- Dependencia para Word ---
try:
    import docx
except ImportError:
    # Esto no detiene el programa si ya tienes la librería, es solo aviso
    pass

# --------- Alta DPI (Windows) ----------
try:
    import ctypes
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

# ---------- helpers portabilidad ----------
def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

BASE_DIR   = app_dir()
DATA_DIR   = BASE_DIR / "data"
ASSETS_DIR = BASE_DIR / "assets"
DATA_DIR.mkdir(exist_ok=True)
ASSETS_DIR.mkdir(exist_ok=True)
DB_PATH    = DATA_DIR / "carreras.db"

# --- BD ---
NOMBRE_BD = str(DB_PATH)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)

# ====== Records por distancia ======
RECORDS = {
    "700":  '700 metros - Record Dist.: 38" 4/5, Sextans 06/04/1997 - Sarfo 01/03/2020',
    "800":  '800 metros - Record Dist.: 43" 2/5, Repirado 20/08/2016',
    "1000": '1.000 metros - Record Dist.: 58" 2/5, Sarfo 23/07/2020',
    "1100": '1.100 metros - Record Dist.: 1\' 04" 3/5, Sold Out 29/08/2021',
    "1200": '1.200 metros - Record Dist.: 1\' 09" 4/5, Donald Music 28/11/2021',
    "1300": '1.300 metros - Record Dist.: 1\' 16" 2/5, Panatta 19/12/2021 - High Commander 08/10/2023 - Jolly Boy 30/11/2025',
    "1400": '1.400 metros - Record Dist.: 1\' 22" 3/5, Patani 29/10/2017 - Dipinto 28/11/2021',
    "1500": '1.500 metros - Record Dist.: 1\' 28" 4/5, Storm Chuck 24/08/2025',
    "1600": '1.600 metros - Record Dist.: 1\' 37" 2/5, Batman Crest 17/08/1975 - Teenek 11/06/2021 - Latan Craf 22/06/2025',
    "1800": '1.800 metros - Record Dist.: 1\' 51" 4/5, Sir Melody 26/11/2017',
    "2000": '2.000 metros - Record Dist.: 2\' 03" 2/5, Volynov 1978',
    "2200": '2.200 metros - Record Dist.: 2\'17"1/5, Frances Net 24/09/2016',
}

# ====== Colores (Hipódromo) ======
COLORS = {
    "bg": "#ffffff",        # blanco
    "primary": "#248689",   # verde azulado
    "accent": "#f16536",    # naranja
    "ink": "#1f2937",
    "muted": "#6b7280",
    "card": "#ffffff",
    "line": "#e5e7eb",
}

# Variable Global para almacenar datos del Word cargado
DATOS_WORD_CACHED = []

# ---------- Utils ----------
def convertir_a_fraccion(valor):
    try:
        valor_str = str(valor).strip().replace('cp', '').strip()
        if '.' in valor_str:
            entero, dec = valor_str.split('.')[0], float('0.' + valor_str.split('.')[1])
            if dec == 0.5: fr = '1/2'
            elif dec == 0.25: fr = '1/4'
            elif dec == 0.75: fr = '3/4'
            else: return valor
            return fr if entero == '0' else f"{entero} {fr}"
        return valor
    except:
        return valor

def _jockey_corto(nombre_jockey: str) -> str:
    s = str(nombre_jockey or "").strip()
    if not s: return ""
    s = re.sub(r"\s+", " ", s)
    if "," in s:
        apellido, resto = [p.strip() for p in s.split(",", 1)]
        nombre = (resto.split()[0] if resto else "")
        return f"{nombre[:1]}. {apellido}".strip() if nombre else apellido
    parts = s.split(" ")
    if len(parts) == 1: return s
    last = parts[-1]
    trailing_is_initial = bool(re.fullmatch(r"[A-Za-z]\.?", last))
    apellido = parts[0]
    nombre = parts[1] if len(parts) >= 2 else parts[0]
    if trailing_is_initial and len(parts) >= 3: nombre = parts[-2]
    inicial = f"{nombre[:1]}." if nombre else ""
    return f"{inicial} {apellido}".strip()

def _july_age_correction(age_value, last_date, today=None):
    if today is None:
        today = date.today()
    try:
        age = int(float(age_value))
    except Exception:
        return age_value
    if pd.isna(last_date):
        return age
    if isinstance(last_date, pd.Timestamp):
        last_date = last_date.date()
    inc = 0
    y = last_date.year
    while True:
        boundary = date(y, 7, 1)
        if last_date < boundary <= today:
            inc += 1
        y += 1
        if date(y, 7, 1) > today:
            break
    return age + inc

def _is_still_debutant(df_acts: pd.DataFrame) -> bool:
    if df_acts is None or df_acts.empty:
        return True
    pf = df_acts.get('Puesto Final', pd.Series([], dtype=object)).astype(str).str.strip().str.upper()
    if not pf.empty and pf.eq('NC').all():
        return True
    obs = df_acts.get('Observacion', pd.Series([], dtype=object)).astype(str).str.lower()
    if (pf.eq('NC') | obs.str.contains('no corr', na=False)).all():
        return True
    return False

def conectar_y_cargar_datos():
    cols_cab = ['Caballo','Edad','Peso','Jockey-Descargo','Padre - Madre','Caballeriza','Cuidador','Pelo']
    cols_act = ['Caballo','Puesto Original','Puesto Final','Jockey','Cuerpos al Ganador',
                'Ganador','Segundo','Margen','Tiempo Ganador','Pista',
                'Fue Distanciado','Fecha','Observacion']
    if not os.path.exists(NOMBRE_BD):
        return pd.DataFrame(columns=cols_cab), pd.DataFrame(columns=cols_act)
    conn = sqlite3.connect(NOMBRE_BD)
    try:
        df_caballos = pd.read_sql_query("SELECT * FROM caballos", conn)
        df_actuaciones = pd.read_sql_query("SELECT * FROM actuaciones", conn)
    finally:
        conn.close()
    df_caballos = df_caballos.rename(columns={
        'nombre': 'Caballo', 'ultima_edad': 'Edad', 'ultimo_peso': 'Peso',
        'ultimo_jockey': 'Jockey-Descargo', 'padre_madre': 'Padre - Madre',
        'caballeriza': 'Caballeriza', 'cuidador': 'Cuidador', 'pelo': 'Pelo'
    })
    df_actuaciones = df_actuaciones.rename(columns={
        'nombre_caballo': 'Caballo', 'puesto_original': 'Puesto Original',
        'puesto_final': 'Puesto Final', 'jockey': 'Jockey',
        'cuerpos': 'Cuerpos al Ganador', 'ganador': 'Ganador',
        'segundo': 'Segundo', 'margen': 'Margen',
        'tiempo_ganador': 'Tiempo Ganador', 'pista': 'Pista',
        'fue_distanciado': 'Fue Distanciado', 'fecha': 'Fecha',
        'observacion': 'Observacion'
    })
    if not df_actuaciones.empty:
        df_actuaciones['Fecha'] = pd.to_datetime(df_actuaciones['Fecha'], errors='coerce')
    for col in ['Caballo','Padre - Madre','Caballeriza','Cuidador','Jockey-Descargo','Pelo','Edad','Peso']:
        if col not in df_caballos.columns: df_caballos[col] = ''
    for col in cols_act:
        if col not in df_actuaciones.columns: df_actuaciones[col] = ''
    if 'Caballo' in df_caballos.columns:
        df_caballos['Caballo'] = df_caballos['Caballo'].astype(str).str.upper().str.strip()
    for col in ['Caballo','Ganador','Segundo']:
        if col in df_actuaciones.columns:
            df_actuaciones[col] = df_actuaciones[col].astype(str).str.upper().str.strip()
    return df_caballos[cols_cab], df_actuaciones[cols_act]

def obtener_datos_caballo(nombre_caballo, db_caballos, db_actuaciones):
    nombre_caballo_upper = nombre_caballo.strip().upper()
    try:
        info_caballo_raw = db_caballos[db_caballos['Caballo'] == nombre_caballo_upper].iloc[0]
    except IndexError:
        return {'Caballo': nombre_caballo_upper, '4 Ult.':'','Nº':'',
                'Pelo':'','Jockey-Descargo':'','E Kg':'','Padre - Madre':'',
                'Caballeriza':'','Cuidador':'','actuaciones': pd.DataFrame()}

    actuaciones_caballo = db_actuaciones[db_actuaciones['Caballo'] == nombre_caballo_upper] \
        .sort_values(by='Fecha', ascending=False)

    debutante = _is_still_debutant(actuaciones_caballo)

    if debutante:
        cuatro = "Debuta"
    else:
        ult=[]
        for p in actuaciones_caballo['Puesto Final'].head(4):
            ps=str(p).strip()
            try:
                n=int(float(ps)); ult.append('0' if n>=10 else str(n))
            except: ult.append(ps if ps else '-')
        cuatro="-".join(reversed(ult)) if ult else ""

    info = info_caballo_raw.copy()

    last_date = actuaciones_caballo['Fecha'].max() if not actuaciones_caballo.empty else pd.NaT
    try:
        edad_ajustada = _july_age_correction(info.get('Edad', ''), last_date)
        peso_val = int(float(info.get('Peso',''))) if str(info.get('Peso','')).strip() != '' else ''
        info['E Kg'] = f"{int(edad_ajustada)} {peso_val}".strip()
    except Exception:
        info['E Kg'] = f"{info.get('Edad','')} {info.get('Peso','')}"

    info['4 Ult.'] = cuatro
    return {**info, 'Nº':'', 'actuaciones': actuaciones_caballo.head(2)}

def _es_yunta(nombre:str)->bool: return bool(re.search(r'\(\s*a\s*\)\s*$', nombre, flags=re.I))
def _limpiar_marca_yunta(nombre:str)->str: return re.sub(r'\(\s*a\s*\)\s*$', '', nombre, flags=re.I).strip()

# ============ LÓGICA WORD (CORREGIDA) ============
def cargar_word_entrada():
    """Lee el Word, parsea carreras solo con encabezados válidos"""
    filepath = filedialog.askopenfilename(filetypes=[("Word Document", "*.docx")])
    if not filepath: return

    try:
        doc = docx.Document(filepath)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el archivo Word.\n{e}")
        return

    global DATOS_WORD_CACHED
    DATOS_WORD_CACHED = []
    
    current_race = {}
    capturing_condition = False
    
    # Palabras clave de inicio (ELIMINADO "PREMIO" PARA EVITAR FALSOS POSITIVOS)
    KEYWORDS_INICIO = ("TURNO", "CLASICO", "CLÁSICO", "ESPECIAL", "GRAN PREMIO", "HANDICAP")
    
    for para in doc.paragraphs:
        texto = para.text.strip()
        if not texto: continue

        # 1. Detectar Cabecera (Inicio de Carrera)
        # Debe empezar EXACTAMENTE con una de las palabras clave
        es_inicio = False
        upper_text = texto.upper()
        for k in KEYWORDS_INICIO:
            # Check: empieza con la palabra Y (es el fin de la linea O hay un espacio después)
            # Esto evita que "TURNO" detecte "TURNOMETRO" (ejemplo tonto) pero asegura firmeza
            if upper_text.startswith(k):
                es_inicio = True
                break
        
        if es_inicio:
            if current_race:
                DATOS_WORD_CACHED.append(current_race)
            
            current_race = {
                "nombre": texto, 
                "distancia": "",
                "condicion_raw": "",
                "premios": ""
            }
            capturing_condition = True
            
            # Extraer distancia de cabecera
            match_dist = re.search(r'(\d{1,2}\.?\d{0,3})\s*(?:metros|mts|m\b)', texto, re.IGNORECASE)
            if match_dist:
                current_race["distancia"] = match_dist.group(1)
            
            continue
        
        # 2. Detectar Premios
        if texto.upper().startswith("PREMIOS:"):
            if current_race:
                vals = texto.split(':', 1)[1].strip()
                current_race["premios"] = vals
            capturing_condition = False
            continue

        # 3. Capturar Cuerpo (Condición)
        if capturing_condition and current_race:
            match_dist_solitaria = re.search(r'^(\d{1,2}\.?\d{0,3})\s*(?:metros|mts|m\b)', texto, re.IGNORECASE)
            if match_dist_solitaria and not current_race["distancia"]:
                current_race["distancia"] = match_dist_solitaria.group(1)
            
            # Solo agregar si no es una linea corta de distancia
            if not (match_dist_solitaria and len(texto) < 20):
                if current_race["condicion_raw"]:
                    current_race["condicion_raw"] += " " + texto
                else:
                    current_race["condicion_raw"] = texto

    if current_race:
        DATOS_WORD_CACHED.append(current_race)
    
    nombres_carreras = [c.get("nombre", "Sin Nombre") for c in DATOS_WORD_CACHED]
    combo_word['values'] = nombres_carreras
    if nombres_carreras:
        combo_word.current(0)
        messagebox.showinfo("Cargado", f"Se detectaron {len(nombres_carreras)} carreras.")
    else:
        messagebox.showwarning("Atención", "No se encontraron carreras con los encabezados: TURNO, CLÁSICO, ESPECIAL...")

def aplicar_seleccion_word(event):
    idx = combo_word.current()
    if idx < 0 or idx >= len(DATOS_WORD_CACHED): return
    
    dato = DATOS_WORD_CACHED[idx]
    
    # 1. Distancia
    dist_str = dato.get("distancia", "")
    dist_limpia = dist_str.replace('.', '').strip()
    
    if dist_limpia in RECORDS:
        combo_dist.set(dist_limpia)
        _aplicar_record()
    else:
        entry_distancia.delete(0, tk.END)
        entry_distancia.insert(0, f"{dist_str} metros")

    # 2. Calcular Categoría y Texto Mágico para el PREMIO
    premios_raw = dato.get("premios", "")
    try:
        d_val = int(dist_limpia)
    except:
        d_val = 0 
    
    # Lógica de Categoría
    cat_txt = "CAT. INTERIOR" # Default (>= 900)
    if d_val <= 800 and d_val > 0:
        cat_txt = "CAT. EXTRAOFICIAL"
    
    # FORMATO: NO COMPUTABLE - CAT. [X] - Premios: $ ...
    linea_premios_completa = f"NO COMPUTABLE - {cat_txt} - Premios: {premios_raw}"

    # 3. Poner en el campo PREMIOS (Como pidió el usuario)
    entry_premios_dinero.delete(0, tk.END)
    entry_premios_dinero.insert(0, linea_premios_completa)
    
    # 4. Poner Condición (Solo Texto Descriptivo)
    condicion_base = dato.get("condicion_raw", "")
    entry_condicion.delete(0, tk.END)
    entry_condicion.insert(0, condicion_base)

# =====================================

def generar_programa_en_tabla():
    if db_caballos.empty:
        messagebox.showwarning("BD vacía",
            "No hay datos en la base.\nCorré 'migracion.py' por consola para crear 'data/carreras.db'.")
        return
    for i in tabla_programa.get_children(): tabla_programa.delete(i)
    text_actuaciones.delete("1.0", tk.END)
    nombres = [n.strip() for n in text_caballos.get("1.0", tk.END).strip().split('\n') if n.strip()]
    numero = 1; ultimo = None
    for nombre_in in nombres:
        es_a=_es_yunta(nombre_in); nombre=_limpiar_marca_yunta(nombre_in)
        datos=obtener_datos_caballo(nombre, db_caballos, db_actuaciones)
        if es_a and ultimo is not None: nro=f"{ultimo}a"
        else: nro=str(numero); ultimo=numero
        if not es_a: numero+=1
        datos['Jockey-Descargo'] = "" 

        datos['Nº'] = nro
        datos['Caballo'] = nombre.upper()
        tabla_programa.insert('', tk.END, values=[datos.get(c, '') for c in columnas])

        acts = datos['actuaciones']
        lineas = []
        debutante = _is_still_debutant(acts)

        if debutante:
            lineas.append("Debutante")
        elif not acts.empty:
            for _, a in acts.sort_values('Fecha').iterrows():
                fecha = a['Fecha'].strftime('%d/%m/%y') if pd.notna(a['Fecha']) else ''
                if str(a['Puesto Final']).strip().upper() == 'NC':
                    obs = str(a.get('Observacion','')).strip()
                    suf = f" ({obs})" if obs else ""
                    lineas.append(f"{fecha} - No Corrió{suf}.")
                    continue

                jockey_corto = _jockey_corto(a.get('Jockey',''))
                dist_txt = " - Distanciado" if str(a.get('Puesto Original','')) != str(a.get('Puesto Final','')) else ""
                puesto_original_str = (
                    f"{int(float(a['Puesto Original']))}º"
                    if str(a.get('Puesto Original','')).strip().replace('.0','').isdigit()
                    else str(a.get('Puesto Original','')).strip()
                )

                if str(a.get('Puesto Original','')).strip() in ('1','1.0'):
                    margen = convertir_a_fraccion(a.get('Margen',''))
                    segundo = str(a.get('Segundo','')).upper().strip()
                    seg_info = db_actuaciones[(db_actuaciones['Caballo'] == segundo) &
                                              (db_actuaciones['Fecha'] == a['Fecha'])]
                    segundo_txt = segundo.title()
                    if not seg_info.empty and str(seg_info.iloc[0].get('Puesto Original','')) != str(seg_info.iloc[0].get('Puesto Final','')):
                        segundo_txt += " (Dist.)"
                    lineas.append(
                        f"{fecha} - {jockey_corto} - 1º gan x {margen} cp a {segundo_txt} - {a.get('Tiempo Ganador','')} - {a.get('Pista','')}{dist_txt}"
                    )
                else:
                    cuerpos_valor = a.get('Cuerpos al Ganador','')
                    terminos_sin_cp = ['S.A','Cza','Pzo','Hco']
                    no_cp = any(t in str(cuerpos_valor) for t in terminos_sin_cp)
                    cuerpos_fmt = convertir_a_fraccion(cuerpos_valor)
                    dif_txt = f"{cuerpos_fmt}" if no_cp else f"{cuerpos_fmt} cp"
                    lineas.append(
                        f"{fecha} - {jockey_corto} - {puesto_original_str} a {dif_txt} de {str(a.get('Ganador','')).title()} - {a.get('Tiempo Ganador','')} - {a.get('Pista','')}{dist_txt}"
                    )

        ant=lineas[0] if len(lineas)>0 else ""; rec=lineas[1] if len(lineas)>1 else ""
        text_actuaciones.insert(tk.END, f"{nro:<2} {ant:<80} | {rec}\n")

programa_completo=[]

def _snapshot_tabla(): return [tabla_programa.item(i)['values'] for i in tabla_programa.get_children()]

def _limpiar_form():
    for entry in [entry_nro_carrera, entry_premio, entry_horario, entry_distancia,
                  entry_condicion, entry_premios_dinero, entry_apuesta, entry_incremento, entry_incremento_2]:
        entry.delete(0, tk.END)
    text_caballos.delete("1.0", tk.END)
    for i in tabla_programa.get_children(): tabla_programa.delete(i)
    text_actuaciones.delete("1.0", tk.END)

def _refrescar_lista_carreras():
    lista_carreras.delete(0, tk.END)
    for carr in programa_completo:
        cab=carr['cabecera']; lista_carreras.insert(tk.END, f"{cab.get('nro_carrera','?')}º - {cab.get('premio','(sin premio)')}")
    contador_carreras.set(f"Carreras Añadidas: {len(programa_completo)}")

def anadir_carrera():
    if not tabla_programa.get_children():
        messagebox.showwarning("Aviso","Primero genera la tabla antes de añadir."); return
    datos={"nro_carrera":entry_nro_carrera.get(), "premio":entry_premio.get(), "horario":entry_horario.get(),
           "distancia":entry_distancia.get(), "condicion":entry_condicion.get(), "premios_dinero":entry_premios_dinero.get(),
           "apuesta":entry_apuesta.get(), "incremento":entry_incremento.get(), "incremento_2":entry_incremento_2.get()}
    programa_completo.append({"cabecera":datos, "tabla_caballos":_snapshot_tabla(), "actuaciones":text_actuaciones.get("1.0", tk.END).strip()})
    _limpiar_form(); _refrescar_lista_carreras(); messagebox.showinfo("OK","Carrera añadida al programa.")

def _cargar_carrera_en_form(c):
    _limpiar_form(); cab=c['cabecera']
    entry_nro_carrera.insert(0,cab.get('nro_carrera','')); entry_premio.insert(0,cab.get('premio',''))
    entry_horario.insert(0,cab.get('horario','')); entry_distancia.insert(0,cab.get('distancia',''))
    entry_condicion.insert(0,cab.get('condicion','')); entry_premios_dinero.insert(0,cab.get('premios_dinero',''))
    entry_apuesta.insert(0,cab.get('apuesta','')); entry_incremento.insert(0,cab.get('incremento',''))
    entry_incremento_2.insert(0,cab.get('incremento_2',''))
    for fila in c['tabla_caballos']: tabla_programa.insert('', tk.END, values=fila)
    text_actuaciones.insert(tk.END, c['actuaciones'])

def editar_carrera():
    sel = lista_carreras.curselection()
    if not sel:
        messagebox.showinfo("Editar", "Seleccioná una carrera del listado.")
        return
    idx = int(sel[0])
    _cargar_carrera_en_form(programa_completo[idx])
    btn_guardar_cambios.config(state=tk.NORMAL)
    btn_eliminar.config(state=tk.NORMAL)
    lista_carreras.config(state=tk.DISABLED)
    btn_anadir.config(state=tk.DISABLED)

def guardar_cambios():
    sel = lista_carreras.curselection()
    if not sel:
        messagebox.showwarning("Aviso", "No hay carrera seleccionada.")
        return
    if not tabla_programa.get_children():
        messagebox.showwarning("Aviso", "Generá/actualizá la tabla antes de guardar.")
        return

    idx = int(sel[0])
    carrera = {"cabecera":{"nro_carrera":entry_nro_carrera.get(),"premio":entry_premio.get(),"horario":entry_horario.get(),
                           "distancia":entry_distancia.get(),"condicion":entry_condicion.get(),"premios_dinero":entry_premios_dinero.get(),
                           "apuesta":entry_apuesta.get(),"incremento":entry_incremento.get(),"incremento_2":entry_incremento_2.get()},
               "tabla_caballos":_snapshot_tabla(),"actuaciones":text_actuaciones.get("1.0", tk.END).strip()}
    programa_completo[idx]=carrera

    _limpiar_form()
    btn_guardar_cambios.config(state=tk.DISABLED)
    btn_eliminar.config(state=tk.DISABLED)
    lista_carreras.config(state=tk.NORMAL)
    btn_anadir.config(state=tk.NORMAL)
    _refrescar_lista_carreras()
    messagebox.showinfo("OK","Cambios guardados.")

def eliminar_carrera():
    sel = lista_carreras.curselection()
    if not sel:
        lista_carreras.config(state=tk.NORMAL)
        btn_anadir.config(state=tk.NORMAL)
        return
    idx = int(sel[0])
    if not (0 <= idx < len(programa_completo)):
        lista_carreras.config(state=tk.NORMAL)
        btn_anadir.config(state=tk.NORMAL)
        return
    if messagebox.askyesno("Eliminar","¿Eliminar la carrera seleccionada?"):
        programa_completo.pop(idx)
        _limpiar_form()
        btn_guardar_cambios.config(state=tk.DISABLED)
        btn_eliminar.config(state=tk.DISABLED)
        lista_carreras.config(state=tk.NORMAL)
        btn_anadir.config(state=tk.NORMAL)
        _refrescar_lista_carreras()

def exportar_a_excel():
    try:
        if not programa_completo:
            messagebox.showwarning("Aviso","No añadiste carreras para exportar."); return
        filepath=filedialog.asksaveasfilename(defaultextension=".xlsx",
                    filetypes=[("Excel Workbook","*.xlsx")], initialdir=str(app_dir()))
        if not filepath: return
        wb=Workbook(); ws=wb.active; ws.title="Programa de Carreras"
        ws.page_margins=PageMargins(left=0.25,right=0.25,top=0.75,bottom=0.75)
        ws.print_options.horizontalCentered=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
        fila=1
        f_tit=Font(name='Tahoma',size=15,bold=True); f_dist=Font(name='Utsaah',size=9,bold=True)
        f_cond=Font(name='Utsaah',size=7); f_pa=Font(name='Arial Narrow',size=8,bold=True)
        f_box=Font(name='Arial Narrow',size=11,bold=True,color="FFFFFF"); fill_negro=PatternFill("solid","000000")
        for c in programa_completo:
            ws.merge_cells(f'C{fila}:I{fila}'); ws[f'C{fila}'].value=c['cabecera']['premio'].upper()
            ws[f'C{fila}'].font=f_tit; ws[f'C{fila}'].alignment=Alignment(horizontal='center',vertical='center')
            ws.merge_cells(f'A{fila}:B{fila}'); cell_nro=ws.cell(row=fila,column=1,value=f"{c['cabecera']['nro_carrera']}º Carrera")
            cell_nro.fill=fill_negro; cell_nro.font=f_box; cell_nro.alignment=Alignment(horizontal='center',vertical='center')
            ws.cell(row=fila,column=10,value=c['cabecera']['horario']).font=f_box
            ws.cell(row=fila,column=10).fill=fill_negro; ws.cell(row=fila,column=10).alignment=Alignment(horizontal='center',vertical='center')
            fila+=1
            ws.merge_cells(f'A{fila}:J{fila}'); ws[f'A{fila}'].value=c['cabecera']['distancia']
            ws[f'A{fila}'].font=f_dist; ws[f'A{fila}'].alignment=Alignment(horizontal='center',vertical='center'); fila+=1
            condicion=c['cabecera']['condicion']; lineas=[x.strip() for x in condicion.split('|')] or [""]
            for lin in lineas:
                ws.merge_cells(f'A{fila}:J{fila}'); ws[f'A{fila}'].value=lin; ws[f'A{fila}'].font=f_cond
                ws[f'A{fila}'].alignment=Alignment(horizontal='left',vertical='center'); fila+=1
            ws.merge_cells(f'A{fila}:H{fila}'); ws[f'A{fila}'].value=c['cabecera']['premios_dinero']; ws[f'A{fila}'].font=f_pa
            if c['cabecera']['incremento']:
                ws.merge_cells(f'I{fila}:J{fila}'); ci=ws[f'I{fila}']; ci.value=c['cabecera']['incremento']
                ci.font=Font(name='Arial Black',size=9,bold=True,italic=True); ci.alignment=Alignment(horizontal='center',vertical='center')
            fila+=1
            ws.merge_cells(f'A{fila}:H{fila}'); ws[f'A{fila}'].value=c['cabecera']['apuesta']; ws[f'A{fila}'].font=f_pa
            if c['cabecera']['incremento_2']:
                ws.merge_cells(f'I{fila}:J{fila}'); ci2=ws[f'I{fila}']; ci2.value=c['cabecera']['incremento_2']
                ci2.font=Font(name='Arial Black',size=9,bold=True,italic=True); ci2.alignment=Alignment(horizontal='center',vertical='center')
            fila+=1
            fila_inicio_tabla=fila
            headers=['4 Ult.','Nº','Caballo','Pelo','Jockey-Descargo','E Kg','Padre - Madre','','Caballeriza','Cuidador']
            ws.merge_cells(f'G{fila}:H{fila}'); ws.cell(row=fila,column=7).value='Padre - Madre'
            for col,t in enumerate(headers,1):
                if col not in (7,8): ws.cell(row=fila,column=col,value=t)
            fila+=1
            for vals in c['tabla_caballos']:
                ws.merge_cells(f'G{fila}:H{fila}')
                for i in range(6): ws.cell(row=fila,column=i+1,value=vals[i])
                ws.cell(row=fila,column=7,value=vals[6]); ws.cell(row=fila,column=9,value=vals[7]); ws.cell(row=fila,column=10,value=vals[8])
                fila+=1
            fila_inicio_act=fila
            for linea in c['actuaciones'].split('\n'):
                if linea.strip():
                    partes=linea.split('|'); num=partes[0].split()[0]; ant=partes[0][len(num):].strip(); rec=partes[1].strip() if len(partes)>1 else ""
                    num_x=int(re.sub(r'\D','',num) or 0)
                    ws.cell(row=fila,column=1,value=num_x); ws.merge_cells(f'B{fila}:F{fila}'); ws.cell(row=fila,column=2,value=ant)
                    ws.merge_cells(f'G{fila}:J{fila}'); ws.cell(row=fila,column=7,value=rec); fila+=1
            med=Side(style="medium"); thin=Side(style="thin")
            for row in ws.iter_rows(min_row=fila_inicio_tabla,max_row=fila-1,min_col=1,max_col=10):
                for cell in row:
                    ali=cell.alignment.copy(); ali.vertical='center'
                    if cell.row>=fila_inicio_act:
                        cell.font=Font(name='Calibri',size=7)
                        if cell.column==1: cell.font=Font(name='Calibri',size=8,bold=True); ali.horizontal='center'
                        if cell.column in (2,7): ali.horizontal='left'
                    elif cell.row==fila_inicio_tabla:
                        cell.font=Font(name='Calibri',size=8,bold=True); ali.horizontal='center'
                    else:
                        cell.font=Font(name='Calibri',size=8)
                        if cell.column in (2,3): cell.font=Font(name='Calibri',size=8,bold=True)
                        if cell.column==1: ali.horizontal='right'
                        if cell.column in (2,4,6): ali.horizontal='center'
                    cell.alignment=ali
                    b=Border(left=thin,right=thin,top=thin,bottom=thin)
                    if cell.row==fila_inicio_tabla: b.top=med
                    if cell.row==fila-1: b.bottom=med
                    if cell.column==1: b.left=med
                    if cell.column==10: b.right=med
                    if cell.row==fila_inicio_tabla: b.bottom=med; b.right=med
                    if cell.row==fila_inicio_act-1: b.bottom=med
                    if fila_inicio_tabla<cell.row<fila_inicio_act and cell.column<10: b.right=med
                    if cell.column==1 and cell.row>=fila_inicio_act: b.right=med
                    if cell.column==6 and cell.row>=fila_inicio_act: b.right=med
                    cell.border=b
            fila+=1
        for k,w in dict(A=6.1,B=3.9,C=15.6,D=5.1,E=13.7,F=3.7,G=9,H=12.4,I=14.6,J=13.7).items():
            ws.column_dimensions[k].width=w
        wb.save(filepath); messagebox.showinfo("Éxito", f"¡Programa exportado en:\n{filepath}!")
    except Exception as e:
        traceback.print_exc(); messagebox.showerror("Error Crítico", f"Ocurrió un error al exportar:\n\n{e}\n\nRevisá la consola.")

# ----------------- UI -----------------
db_caballos, db_actuaciones = conectar_y_cargar_datos()

root = tk.Tk()
root.title("Programa")
root.configure(bg=COLORS["bg"])

# Icono .ICO real
try:
    ico = ASSETS_DIR / "programa.ico"
    if ico.exists():
        root.iconbitmap(default=str(ico))
except Exception:
    pass

# ----- Estilos -----
style = ttk.Style()
try: style.theme_use('clam')
except Exception: pass
style.configure(".", background=COLORS["bg"], foreground=COLORS["ink"])
style.configure("Card.TFrame", background=COLORS["card"])
style.configure("Top.TFrame",  background=COLORS["card"])
style.configure("TLabel", background=COLORS["card"], foreground=COLORS["ink"])
style.configure("Field.TLabel", background=COLORS["card"], foreground=COLORS["accent"], font=("Segoe UI", 9, "bold"))
style.configure("Muted.TLabel", foreground=COLORS["muted"], background=COLORS["card"])
style.configure("Title.TLabel", font=("Segoe UI", 18, "bold"), foreground=COLORS["primary"], background=COLORS["card"])
style.configure("TButton", background=COLORS["primary"], foreground="white", padding=6, relief="flat")
style.map("TButton", background=[("active", "#1f6f71")])
style.configure("Accent.TButton", background=COLORS["accent"], foreground="white")
style.map("Accent.TButton", background=[("active", "#d7572f")])
style.configure("Treeview", background="white", foreground=COLORS["ink"], fieldbackground="white", rowheight=24,
                bordercolor=COLORS["line"], borderwidth=1)
style.configure("Treeview.Heading", background=COLORS["primary"], foreground="white", font=("Segoe UI", 9, "bold"))

# ----- Header -----
header = ttk.Frame(root, style="Top.TFrame")
header.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(10,6))
logo_lbl = ttk.Label(header, style="Top.TFrame")
logo_lbl.pack(side=tk.LEFT, padx=(0,10))
try:
    _logo_img = tk.PhotoImage(file=str(ASSETS_DIR / "logo_hipodromo.png"))
    logo_lbl.configure(image=_logo_img)
except Exception:
    pass
title_box = ttk.Frame(header, style="Top.TFrame"); title_box.pack(side=tk.LEFT, fill=tk.X, expand=True)
ttk.Label(title_box, text="Programa", style="Title.TLabel").pack(anchor="w")
ttk.Label(title_box, text="Creador de Programas de Carreras", style="Muted.TLabel").pack(anchor="w")

# BOTÓN CARGAR WORD
btn_word = ttk.Button(header, text="📂 Cargar Programa (Word)", command=cargar_word_entrada)
btn_word.pack(side=tk.RIGHT, padx=5)

ttk.Separator(root, orient="horizontal").pack(fill=tk.X, padx=10, pady=(0,8))

# ----- Top form -----
frame_top = ttk.Frame(root, style="Card.TFrame", padding=10)
frame_top.pack(side=tk.TOP, fill=tk.X, padx=10)

def _field(parent, label, entry_w=30):
    ttk.Label(parent, text=label, style="Field.TLabel").pack(anchor="w")
    e = ttk.Entry(parent, width=entry_w); e.pack(fill=tk.X)
    return e

col1 = ttk.Frame(frame_top, style="Card.TFrame"); col1.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

# COMBOBOX WORD
ttk.Label(col1, text="Seleccionar Turno/Carrera (Word):", style="Field.TLabel", foreground=COLORS["primary"]).pack(anchor="w")
combo_word = ttk.Combobox(col1, state="readonly", width=35)
combo_word.pack(fill=tk.X, pady=(0, 10))
combo_word.bind("<<ComboboxSelected>>", aplicar_seleccion_word)

entry_nro_carrera   = _field(col1, "Nº Carrera:")
entry_premio        = _field(col1, "Nombre del Premio:")
entry_horario       = _field(col1, "Horario:")

col2 = ttk.Frame(frame_top, style="Card.TFrame"); col2.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
ttk.Label(col2, text="Distancia:", style="Field.TLabel").pack(anchor="w")
dist_var = tk.StringVar()
combo_dist = ttk.Combobox(col2, textvariable=dist_var, values=list(RECORDS.keys()), state="readonly"); combo_dist.pack(fill=tk.X)
ttk.Label(col2, text="Distancia y Récord (editable):", style="Field.TLabel").pack(anchor="w", pady=(6,0))
entry_distancia     = _field(col2, "", 40); entry_distancia.pack_configure()
ttk.Label(col2, text="Condición (usar '|' para nueva línea):", style="Field.TLabel").pack(anchor="w", pady=(10,0))
entry_condicion     = _field(col2, "", 40)
ttk.Label(col2, text="Premios ($):", style="Field.TLabel").pack(anchor="w", pady=(10,0))
entry_premios_dinero= _field(col2, "", 40)

def _aplicar_record(*_):
    k = dist_var.get().strip()
    if k in RECORDS:
        entry_distancia.delete(0, tk.END)
        entry_distancia.insert(0, RECORDS[k])
combo_dist.bind("<<ComboboxSelected>>", _aplicar_record)

col3 = ttk.Frame(frame_top, style="Card.TFrame"); col3.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
entry_apuesta       = _field(col3, "Apuesta:", 50)
entry_incremento    = _field(col3, "Incremento ($) - (Al lado de Premios):", 50)
entry_incremento_2  = _field(col3, "Incremento 2 ($) - (Al lado de Apuesta):", 50)
ttk.Label(col3, text="Lista de Caballos (uno por línea). Para yunta, agregá '(a)' al segundo.",
          style="Field.TLabel").pack(anchor="w", pady=(10,0))
text_caballos = tk.Text(col3, height=5, width=50, bg="white"); text_caballos.pack(fill=tk.X, expand=True)

# ======= Área media =======
class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = tk.Canvas(self, highlightthickness=0, bg=COLORS["bg"])
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = ttk.Frame(self.canvas, style="Card.TFrame")
        self.window_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.window_id, width=e.width))

        self.canvas.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")

        self.canvas.bind_all("<MouseWheel>", lambda ev: self.canvas.yview_scroll(int(-1*(ev.delta/120)), "units"))

wrap_mid = ScrollableFrame(root)
wrap_mid.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
frame_mid = wrap_mid.inner

frame_right = ttk.Frame(frame_mid, style="Card.TFrame")
frame_right.pack(side=tk.RIGHT, fill=tk.Y, padx=(8, 0))
frame_left = ttk.Frame(frame_mid, style="Card.TFrame")
frame_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 0))

columnas = ('4 Ult.','Nº','Caballo','Pelo','Jockey-Descargo','E Kg','Padre - Madre','Caballeriza','Cuidador')
tabla_programa = ttk.Treeview(frame_left, columns=columnas, show='headings')
for col in columnas:
    tabla_programa.heading(col, text=col)
    ancho = 120
    if col in ('4 Ult.','Nº','Pelo','E Kg'): ancho = 80
    if col in ('Padre - Madre','Jockey-Descargo'): ancho = 165
    tabla_programa.column(col, width=ancho, anchor="w")
tabla_programa.pack(fill=tk.BOTH, expand=True, pady=(5, 5))

ttk.Label(frame_left, text="Últimas Actuaciones Detalladas:", style="Field.TLabel").pack(anchor="w", pady=(4, 0))
text_actuaciones = tk.Text(frame_left, height=10, font=("Courier New", 9), bg="white")
text_actuaciones.pack(fill=tk.BOTH, expand=True)

right_inner = ttk.Frame(frame_right, style="Card.TFrame")
right_inner.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

ttk.Label(right_inner, text="Carreras añadidas", style="Field.TLabel").pack(anchor="w", pady=(0, 6))

list_frame = ttk.Frame(right_inner, style="Card.TFrame")
list_frame.pack(fill=tk.BOTH, expand=True)

lista_scroll = ttk.Scrollbar(list_frame, orient="vertical")
lista_carreras = tk.Listbox(list_frame, yscrollcommand=lista_scroll.set)
lista_scroll.config(command=lista_carreras.yview)

lista_carreras.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
lista_scroll.pack(side=tk.RIGHT, fill=tk.Y)

buttons_right = ttk.Frame(right_inner, style="Card.TFrame")
buttons_right.pack(fill=tk.X, pady=(8, 0))

btn_editar = ttk.Button(buttons_right, text="Editar seleccionada", command=editar_carrera)
btn_editar.pack(fill=tk.X, pady=2)

btn_guardar_cambios = ttk.Button(buttons_right, text="Guardar cambios",
                                 command=guardar_cambios, state=tk.DISABLED)
btn_guardar_cambios.pack(fill=tk.X, pady=2)

btn_eliminar = ttk.Button(buttons_right, text="Eliminar seleccionada",
                          command=eliminar_carrera, state=tk.DISABLED)
btn_eliminar.pack(fill=tk.X, pady=2)

# ======= Botonera inferior =======
frame_bot = ttk.Frame(root, style="Card.TFrame", padding=10)
frame_bot.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 10))

contador_carreras = tk.StringVar(value="Carreras Añadidas: 0")
ttk.Label(frame_bot, textvariable=contador_carreras).pack(side=tk.LEFT)

btn_generar = ttk.Button(frame_bot, text="Generar Programa",
                          command=generar_programa_en_tabla, style="Accent.TButton")
btn_generar.pack(side=tk.RIGHT, padx=5)

btn_anadir = ttk.Button(frame_bot, text="Añadir Carrera", command=anadir_carrera)
btn_anadir.pack(side=tk.RIGHT, padx=5)

btn_exportar = ttk.Button(frame_bot, text="Exportar Programa Completo",
                          command=exportar_a_excel)
btn_exportar.pack(side=tk.RIGHT, padx=5)

# Cargar BD
db_caballos, db_actuaciones = conectar_y_cargar_datos()

# Geometría
sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
w = min(1550, max(1100, sw - 40)); h = min(930, max(650, sh - 80))
root.geometry(f"{w}x{h}+10+10"); root.minsize(1100, 650)

root.mainloop()